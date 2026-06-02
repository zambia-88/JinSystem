#!/usr/bin/env python3
"""将个人知识库 Excel 导出为 source/docs Markdown（权限由 Web 控制公开范围）。"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("请先安装: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT.parent / "个人知识库2025-2035.xlsx"
SOURCE_DOCS = ROOT / "source" / "docs"
POSTS_JSON = SOURCE_DOCS / ".vitepress" / "posts.json"  # 占位，posts 由 build_public 生成

PUBLIC_SHEETS: dict[str, tuple[str, str, str]] = {
    "cognition/mindset": ("格局提升", "格局提升", "修行、长期主义与内心秩序"),
    "cognition/insight": ("人性洞察", "人性洞察", "人际、职场与自我认知"),
    "tools/ai": ("人工智能", "AI 工具库", "实用 AI 工具与资源链接"),
    "tools/excel": ("Excel", "Excel 技巧", "表格处理与效率技巧"),
    "tools/word": ("Word", "Word 技巧", "文档排版与审阅技巧"),
    "tools/ppt": ("PPT", "PPT 技巧", "演示文稿制作与 AI 辅助"),
    "life/daily": ("日常生活", "日常生活", "居家、健康与生活窍门"),
    "life/travel": ("旅游", "旅游攻略", "路线规划与小众目的地"),
}

SKIP_TITLE_KEYWORDS = re.compile(
    r"wifi|密码|破解|regedit|WeChat|微信存储|netsh",
    re.IGNORECASE,
)

MEDIA_TYPES = {"image", "video", "audio", "图片", "视频", "音频", "音乐"}


def slugify(text: str, index: int) -> str:
    text = (text or "").strip().split("\n")[0][:40]
    text = re.sub(r"[^\w\u4e00-\u9fff\-]", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return f"{index:03d}-{text}" if text else f"{index:03d}-entry"


def fmt_date(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).split()[0]
    s = s.replace(".", "-")
    return s


def cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def esc_yaml(value: str) -> str:
    return value.replace('"', "'").replace("\n", " ")


def make_excerpt(text: str, max_len: int = 120) -> str:
    t = re.sub(r"\s+", " ", (text or "")).strip()
    if len(t) <= max_len:
        return t
    return t[: max_len - 1] + "…"


def normalize_media_type(raw: str) -> str:
    mapping = {
        "图片": "image",
        "image": "image",
        "视频": "video",
        "video": "video",
        "音频": "audio",
        "音乐": "audio",
        "audio": "audio",
    }
    return mapping.get(raw.lower() if raw.isascii() else raw, mapping.get(raw, ""))


OFFLINE_VALUES = {"否", "no", "n", "0", "不公开", "下线", "隐藏"}
DELETE_VALUES = {"是", "yes", "y", "1", "删除", "删"}


def flag_match(value: str, options: set[str]) -> bool:
    v = (value or "").strip()
    if not v:
        return False
    lowered = {o.lower() for o in options}
    return v.lower() in lowered


def row_is_published(cells: list[str], get) -> bool:
    """仅尊重「删除」列；公开/非公开由 Web permissions.json 控制。"""
    if flag_match(get(cells, "删除"), DELETE_VALUES):
        return False
    return True


def is_valid_row(title: str, content: str) -> bool:
    if not title and not content:
        return False
    if title in ("序号", "主题/标题"):
        return False
    combined = f"{title} {content}"
    if SKIP_TITLE_KEYWORDS.search(combined):
        return False
    return bool(title or content)


def parse_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    entries: list[dict] = []
    header_idx = None

    for i, row in enumerate(rows):
        cells = [cell_str(c) for c in row]
        if "主题/标题" in cells or "主题" in cells:
            header_idx = i
            break

    if header_idx is None:
        return entries

    header = [cell_str(c) for c in rows[header_idx]]
    col = {name: idx for idx, name in enumerate(header) if name}

    def get(row, *names: str) -> str:
        for name in names:
            if name in col and col[name] < len(row):
                v = cell_str(row[col[name]])
                if v:
                    return v
        return ""

    for row in rows[header_idx + 1 :]:
        cells = [cell_str(c) for c in row]
        if not any(cells):
            continue
        title = get(cells, "主题/标题", "主题")
        content = get(cells, "核心内容简述", "核心内容", "主要事件")
        if not is_valid_row(title, content):
            continue
        if not row_is_published(cells, get):
            continue
        date_val = ""
        if "记录日期" in col and col["记录日期"] < len(row):
            date_val = fmt_date(row[col["记录日期"]])
        elif "时间" in col and col["时间"] < len(row):
            date_val = fmt_date(row[col["时间"]])

        cover = get(cells, "封面", "封面图", "cover")
        media_type = normalize_media_type(get(cells, "媒体类型", "mediaType"))
        media_url = get(cells, "媒体链接", "媒体地址", "mediaUrl", "视频", "音频")

        entries.append(
            {
                "index": get(cells, "序号") or str(len(entries) + 1),
                "title": title or f"条目 {len(entries) + 1}",
                "content": content,
                "source": get(cells, "来源/出处", "来源"),
                "date": date_val,
                "reflection": get(cells, "个人感悟"),
                "note": get(cells, "说明"),
                "cover": cover,
                "mediaType": media_type,
                "mediaUrl": media_url,
            }
        )
    return entries


def read_existing_frontmatter(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    fm, _ = read_frontmatter_file(path.read_text(encoding="utf-8"))
    return fm


def read_frontmatter_file(text: str) -> tuple[dict[str, str], str]:
    if not text.startswith("---"):
        return {}, text
    parts = text.split("---", 2)
    if len(parts) < 3:
        return {}, text
    fm: dict[str, str] = {}
    for line in parts[1].strip().splitlines():
        if ":" in line:
            k, _, v = line.partition(":")
            fm[k.strip()] = v.strip().strip('"').strip("'")
    return fm, parts[2]


def write_entry(path: Path, entry: dict, category: str, slug_path: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = read_existing_frontmatter(path)
    if not entry["cover"] and existing.get("cover"):
        entry["cover"] = existing["cover"]
    if not entry["mediaType"] and existing.get("mediaType"):
        entry["mediaType"] = existing["mediaType"]
    if not entry["mediaUrl"] and existing.get("mediaUrl"):
        entry["mediaUrl"] = existing["mediaUrl"]
    excerpt = make_excerpt(entry["content"])
    lines = [
        "---",
        f'title: "{esc_yaml(entry["title"])}"',
        f"category: {category}",
        f'excerpt: "{esc_yaml(excerpt)}"',
    ]
    if entry["date"]:
        lines.append(f"date: {entry['date']}")
    if entry["source"]:
        lines.append(f'source: "{esc_yaml(entry["source"])}"')
    if entry["cover"]:
        lines.append(f'cover: "{esc_yaml(entry["cover"])}"')
    if entry["mediaType"]:
        lines.append(f'mediaType: {entry["mediaType"]}')
    if entry["mediaUrl"]:
        lines.append(f'mediaUrl: "{esc_yaml(entry["mediaUrl"])}"')
    lines.extend(["---", "", f"# {entry['title']}", ""])
    if entry["content"]:
        lines.append(entry["content"])
        lines.append("")
    if entry["reflection"]:
        lines.append("## 个人感悟")
        lines.append("")
        lines.append(f"> {entry['reflection']}")
        lines.append("")
    if entry["source"]:
        src = entry["source"]
        if src.startswith("http"):
            lines.append(f"**来源：** [{src}]({src})")
        else:
            lines.append(f"**来源：** {src}")
        lines.append("")
    if entry["note"] and entry["note"] != entry["reflection"]:
        lines.append(f"*{entry['note']}*")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_index(path: Path, title: str, desc: str, items: list[tuple[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "---",
        f"title: {title}",
        "---",
        "",
        f"# {title}",
        "",
        desc,
        "",
    ]
    for name, link in items:
        lines.append(f"- [{name}]({link})")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def cleanup_stale_markdown(cat_dir: Path, keep: set[Path]) -> int:
    """删除 Excel 中已不存在或已下线的 .md（保留 index.md）。"""
    removed = 0
    if not cat_dir.is_dir():
        return removed
    for md in cat_dir.glob("*.md"):
        if md.name == "index.md" or md in keep:
            continue
        md.unlink()
        print(f"  删除: {md.relative_to(ROOT)}")
        removed += 1
    return removed


def main() -> None:
    if not EXCEL.exists():
        raise SystemExit(f"找不到 Excel: {EXCEL}")

    from pull_permissions import pull as pull_permissions

    pull_permissions()

    from permissions_lib import (
        ensure_permission_entry,
        load_manifest,
        load_permissions,
        save_manifest,
        save_permissions,
        slug_from_path,
        upsert_manifest_entry,
    )

    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    stats: dict[str, int] = {}
    perms = load_permissions()
    manifest = load_manifest()

    for slug, (sheet_name, cn_title, desc) in PUBLIC_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        entries = parse_sheet(ws)
        cat_dir = SOURCE_DOCS / slug.replace("/", "/")
        items: list[tuple[str, str]] = []

        generated: set[Path] = {cat_dir / "index.md"}
        for i, entry in enumerate(entries, start=1):
            try:
                num = int(re.sub(r"\D", "", str(entry["index"])) or i)
            except ValueError:
                num = i
            fname = slugify(entry["title"], num) + ".md"
            rel = f"{slug}/{fname}"
            link_slug = f"/{slug}/{fname.replace('.md', '')}"
            md_path = cat_dir / fname
            generated.add(md_path)
            existing = read_existing_frontmatter(md_path)
            if not entry["cover"] and existing.get("cover"):
                entry["cover"] = existing["cover"]
            if not entry["mediaType"] and existing.get("mediaType"):
                entry["mediaType"] = existing["mediaType"]
            if not entry["mediaUrl"] and existing.get("mediaUrl"):
                entry["mediaUrl"] = existing["mediaUrl"]
            write_entry(md_path, entry, cn_title, slug)
            items.append((entry["title"].split("\n")[0], link_slug))
            upsert_manifest_entry(
                manifest,
                link_slug,
                entry["title"].split("\n")[0],
                cn_title,
                "excel",
                rel,
            )
            ensure_permission_entry(perms, link_slug, "excel")

        write_index(cat_dir / "index.md", cn_title, desc, items)
        removed = cleanup_stale_markdown(cat_dir, generated)
        stats[slug] = len(entries)
        if removed:
            print(f"  {slug}: 清理 {removed} 个已删条目")

    wb.close()
    save_manifest(manifest)
    save_permissions(perms)

    total = sum(stats.values())
    print("导出到 source/docs 完成:")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    print(f"  合计: {total}")
    print("  下一步: python scripts/build_public.py")


if __name__ == "__main__":
    main()
